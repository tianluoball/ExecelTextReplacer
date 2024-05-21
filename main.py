import os
import glob
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
import pyperclip
import re

def get_excel_files(directory):
    excel_files = glob.glob(os.path.join(directory, "*.xlsx"))
    excel_files += glob.glob(os.path.join(directory, "*.xls"))
    return excel_files

def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")])
    return file_path

def check_file_type(file_path):
    if file_path.lower().endswith('.numbers'):
        messagebox.showinfo(translation["reminder"][language], translation["convert_numbers"][language])
        return None
    return file_path

def on_run_click():
    default_text = input_text_box.get("1.0", tk.END).strip()
    arabic_number = arabic_number_var.get()
    if default_text and arabic_number:
        replaced_text = replace_special_strings(default_text, arabic_number, selected_file)
        output_text_box.delete("1.0", tk.END)
        output_text_box.insert(tk.END, replaced_text)
        pyperclip.copy(replaced_text)  # Copy replaced text to clipboard
        arabic_number_var.set(arabic_number + 1)
    else:
        messagebox.showwarning(translation["warning"][language], translation["text_and_number_warning"][language])

def replace_special_strings(text, row_number, excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    
    def replace_match(match):
        col_label = match.group(1)
        col_index = ord(col_label) - ord('A') + 1  # openpyxl column index starts from 1
        cell_value = ws.cell(row=row_number, column=col_index).value
        return str(cell_value) if cell_value is not None else ''
    
    pattern = re.compile(r'&([A-Z]+)&')
    replaced_text = pattern.sub(replace_match, text)
    return replaced_text

def select_and_check_file():
    global selected_file
    selected_file = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
    if selected_file:
        selected_file = check_file_type(selected_file)
        if selected_file:
            messagebox.showinfo(translation["reminder"][language], translation["file_selected"][language])
        else:
            messagebox.showwarning(translation["reminder"][language], translation["select_excel"][language])

def toggle_language():
    global language
    language = "en" if language == "zh" else "zh"
    update_texts()

def update_texts():
    input_text_box_label.config(text=translation["input_text"][language])
    instruction_label.config(text=translation["instructions"][language])
    select_file_button.config(text=translation["select_file"][language])
    run_button.config(text=translation["run"][language])
    switch_language_button.config(text=translation["switch_language"][language])

# Translations dictionary
translation = {
    "reminder": {"en": "Reminder", "zh": "提醒"},
    "convert_numbers": {"en": "Please convert Numbers file to Excel format before proceeding.", "zh": "请将 Numbers 文件转换为 Excel 文件后再继续操作。"},
    "warning": {"en": "Warning", "zh": "警告"},
    "text_and_number_warning": {"en": "Text box and number input cannot be empty!", "zh": "文本框和数字输入框不能为空！"},
    "input_text": {"en": "Input Text", "zh": "输入文本"},
    "instructions": {"en": "Use &letter& format to specify columns to replace. Enter a number to specify Excel row number.", "zh": "使用 &字母& 的形式来确定要替换文本所处的列。输入数字来确定Excel行数。"},
    "select_file": {"en": "Select File", "zh": "选择文件"},
    "run": {"en": "Run", "zh": "运行"},
    "switch_language": {"en": "Switch to Chinese", "zh": "切换到英文"},
    "file_selected": {"en": "File selected successfully!", "zh": "文件选择成功！"},
    "select_excel": {"en": "Please select an Excel file.", "zh": "请选择一个 Excel 文件。"}
}

# Initial language
language = "en"

# Create main window
root = tk.Tk()
root.title("Text Replacement Tool")

# Create input text label and text box
input_text_box_label = tk.Label(root, text=translation["input_text"][language])
input_text_box_label.pack(pady=5)
input_text_box = tk.Text(root, height=10, width=50)
input_text_box.pack(pady=10)

# Add instruction label
instruction_label = tk.Label(root, text=translation["instructions"][language])
instruction_label.pack(pady=5)

# Create select file button
select_file_button = tk.Button(root, text=translation["select_file"][language], command=select_and_check_file)
select_file_button.pack(pady=5)

# Create Arabic number input
arabic_number_var = tk.IntVar(value=1)
arabic_number_entry = tk.Entry(root, textvariable=arabic_number_var)
arabic_number_entry.pack(pady=5)

# Create run button
run_button = tk.Button(root, text=translation["run"][language], command=on_run_click)
run_button.pack(pady=5)

# Create output text box
output_text_box = tk.Text(root, height=10, width=50, bg="lightgrey")
output_text_box.pack(pady=10)

# Create switch language button
switch_language_button = tk.Button(root, text=translation["switch_language"][language], command=toggle_language)
switch_language_button.pack(pady=5)

# Run main loop
root.mainloop()
